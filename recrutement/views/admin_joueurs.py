import io
import os

from django.conf import settings
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone

from effectif.forms import JoueurForm
from effectif.models import Joueur

from .admin_utils import base_ctx, staff_required


@staff_required
def admin_joueurs(request):
    from django.db.models import Count
    q = request.GET.get('q', '')
    poste = request.GET.get('poste', 'all')
    qs = Joueur.objects.all()
    if q:
        qs = qs.filter(nom__icontains=q) | qs.filter(prenom__icontains=q)
    if poste != 'all':
        qs = qs.filter(poste=poste)
    stats = {item['poste']: item['n'] for item in
             Joueur.objects.values('poste').annotate(n=Count('pk'))}
    ctx = base_ctx(request, 'joueurs')
    ctx.update({'joueurs': qs, 'q': q, 'poste_actif': poste,
                'total': Joueur.objects.count(), 'stats_postes': stats})
    return render(request, 'admin_cfld/joueurs.html', ctx)


@staff_required
def admin_joueurs_excel(request):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    q = request.GET.get('q', '')
    poste = request.GET.get('poste', 'all')
    cat = request.GET.get('categorie', 'all')
    actif = request.GET.get('actif', 'all')

    qs = Joueur.objects.select_related('candidature').order_by('categorie', 'sexe', 'poste', 'numero')
    if q:
        qs = qs.filter(nom__icontains=q) | qs.filter(prenom__icontains=q)
    if poste != 'all':
        qs = qs.filter(poste=poste)
    if cat != 'all':
        qs = qs.filter(categorie=cat)
    if actif == 'actif':
        qs = qs.filter(actif=True)
    elif actif == 'inactif':
        qs = qs.filter(actif=False)

    wb = Workbook()
    ws = wb.active
    ws.title = "Joueurs CFLD"

    rouge = "C8161D"
    blanc = "FFFFFF"
    gris_ligne = "EDEBE4"

    hd_font = Font(bold=True, color=blanc, size=10, name="Calibri")
    hd_fill = PatternFill("solid", fgColor=rouge)
    hd_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="D1D5DB")
    thin_bord = Border(left=thin, right=thin, top=thin, bottom=thin)
    alt_fill = PatternFill("solid", fgColor=gris_ligne)
    body_font = Font(size=9, name="Calibri")
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")

    headers = [
        "Réf. CFLD", "N°", "Nom", "Prénom", "Catégorie", "Poste", "Sexe",
        "Date naissance", "Âge", "Nationalité", "Pied fort",
        "Téléphone WA", "Email", "Adresse",
        "Nom parent", "Tél. parent", "Email parent",
        "Ancien club", "N° licence", "Date inscription",
        "Matchs", "Buts", "Info scolaire", "Contact urgence", "Actif",
    ]
    ws.append(headers)
    for col_idx, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = hd_font
        cell.fill = hd_fill
        cell.alignment = hd_align
        cell.border = thin_bord
    ws.row_dimensions[1].height = 28

    for i, j in enumerate(qs, 2):
        row = [
            j.candidature.reference if j.candidature else '',
            j.numero, j.nom, j.prenom, j.categorie,
            j.get_poste_display(), j.get_sexe_display(),
            j.date_naissance.strftime('%d/%m/%Y') if j.date_naissance else '',
            j.age, j.nationalite, j.pied_fort or '',
            j.telephone_whatsapp or '', j.email or '', j.adresse or '',
            j.nom_parent or '', j.telephone_parent or '', j.email_parent or '',
            j.ancien_club or '', j.numero_licence or '',
            j.date_inscription.strftime('%d/%m/%Y') if j.date_inscription else '',
            j.matchs_joues, j.buts, j.info_scolaire or '',
            j.contact_urgence or '', 'Oui' if j.actif else 'Non',
        ]
        ws.append(row)
        fill = alt_fill if i % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
        for col_idx, _ in enumerate(row, 1):
            cell = ws.cell(row=i, column=col_idx)
            cell.font = body_font
            cell.fill = fill
            cell.border = thin_bord
            cell.alignment = center if col_idx in (1, 4, 5, 6, 8, 10, 20, 21, 24) else left
        ws.row_dimensions[i].height = 18

    col_widths = [14, 6, 18, 18, 10, 14, 10, 14, 6, 16, 12,
                  16, 26, 22, 20, 16, 26, 18, 14, 16, 8, 6, 22, 22, 7]
    for idx, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(idx)].width = w

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    date_str = timezone.localdate().strftime('%Y%m%d')
    response = HttpResponse(
        buf.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="joueurs_cfld_{date_str}.xlsx"'
    return response


@staff_required
def admin_joueur_form(request, pk=None):
    obj = get_object_or_404(Joueur, pk=pk) if pk else None
    form = JoueurForm(request.POST or None, request.FILES or None, instance=obj)
    if request.method == 'POST' and form.is_valid():
        form.save()
        return redirect('admin_joueurs')
    ctx = base_ctx(request, 'joueurs')
    ctx.update({'form': form, 'obj': obj})
    return render(request, 'admin_cfld/joueur_form.html', ctx)


@staff_required
def admin_joueur_detail(request, pk):
    joueur = get_object_or_404(Joueur, pk=pk)
    ctx = base_ctx(request, 'joueurs')
    ctx.update({'joueur': joueur})
    return render(request, 'admin_cfld/joueur_detail.html', ctx)


@staff_required
def admin_joueur_delete(request, pk):
    if request.method != 'POST':
        return JsonResponse({'error': 'POST requis'}, status=405)
    get_object_or_404(Joueur, pk=pk).delete()
    return JsonResponse({'ok': True})


@staff_required
def admin_joueur_pdf(request, pk):
    from PIL import Image as PILImage
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import (
        BaseDocTemplate, Frame, Image as RLImage, KeepTogether,
        PageTemplate, Paragraph, Spacer, Table, TableStyle,
    )

    joueur = get_object_or_404(Joueur, pk=pk)

    C_ROUGE = colors.HexColor('#C8161D')
    C_ROUGE_DARK = colors.HexColor('#8A0D12')
    C_ROUGE_LIGHT = colors.HexColor('#FFF0F0')
    C_NOIR = colors.HexColor('#0B0B0B')
    C_GRIS = colors.HexColor('#6B7280')
    C_GRIS_LIGHT = colors.HexColor('#F5F5F5')
    C_GRIS_LINE = colors.HexColor('#E5E7EB')
    C_BLANC = colors.white

    W, H = A4
    MARGIN_L = 1.8 * cm
    MARGIN_R = 1.8 * cm
    MARGIN_T = 3.8 * cm
    MARGIN_B = 2.0 * cm
    CONTENT_W = W - MARGIN_L - MARGIN_R

    logo_path = os.path.join(settings.BASE_DIR, 'assets', 'images', 'club', 'logo.png')

    def draw_header(canvas, doc):
        canvas.saveState()
        HEADER_H = 3.0 * cm
        canvas.setFillColor(C_ROUGE_DARK)
        canvas.rect(0, H - HEADER_H, W, HEADER_H, stroke=0, fill=1)
        canvas.setFillColor(C_ROUGE)
        canvas.rect(0, H - HEADER_H, 0.35 * cm, HEADER_H, stroke=0, fill=1)

        if os.path.exists(logo_path):
            logo_size = 2.0 * cm
            logo_x = 0.7 * cm
            logo_y = H - HEADER_H + (HEADER_H - logo_size) / 2
            try:
                canvas.drawImage(logo_path, logo_x, logo_y,
                                 width=logo_size, height=logo_size,
                                 preserveAspectRatio=True, mask='auto')
            except Exception:
                pass
            title_x = logo_x + logo_size + 0.5 * cm
        else:
            title_x = 0.7 * cm

        canvas.setFillColor(C_BLANC)
        canvas.setFont('Helvetica-Bold', 16)
        canvas.drawString(title_x, H - 1.2 * cm, 'FICHE JOUEUR')
        canvas.setFillColor(colors.HexColor('#FFBCBC'))
        canvas.setFont('Helvetica', 8.5)
        canvas.drawString(title_x, H - 1.85 * cm,
                          "Centre de Formation Lancine Diomandé  ·  CFLD  ·  Aboisso, Côte d'Ivoire")

        badge_w = 2.8 * cm
        badge_x = W - badge_w - 0.6 * cm
        badge_y = H - 1.1 * cm
        canvas.setFillColor(C_ROUGE)
        canvas.roundRect(badge_x, badge_y - 0.7 * cm, badge_w, 0.7 * cm, 4, stroke=0, fill=1)
        canvas.setFillColor(C_BLANC)
        canvas.setFont('Helvetica-Bold', 10)
        num_text = f'N° {joueur.numero}'
        nw = canvas.stringWidth(num_text, 'Helvetica-Bold', 10)
        canvas.drawString(badge_x + (badge_w - nw) / 2, badge_y - 0.48 * cm, num_text)

        canvas.setFillColor(colors.HexColor('#FFBCBC'))
        canvas.setFont('Helvetica', 7.5)
        sec = joueur.section
        sw = canvas.stringWidth(sec, 'Helvetica', 7.5)
        canvas.drawString(badge_x + (badge_w - sw) / 2, badge_y - 1.0 * cm, sec)

        canvas.setStrokeColor(C_ROUGE)
        canvas.setLineWidth(1)
        canvas.line(0, H - HEADER_H - 0.4, W, H - HEADER_H - 0.4)
        canvas.restoreState()

    def draw_footer(canvas, doc):
        canvas.saveState()
        y = MARGIN_B - 0.4 * cm
        canvas.setStrokeColor(C_GRIS_LINE)
        canvas.setLineWidth(0.5)
        canvas.line(MARGIN_L, y + 0.9 * cm, W - MARGIN_R, y + 0.9 * cm)
        canvas.setFont('Helvetica', 7)
        canvas.setFillColor(C_GRIS)
        date_str = timezone.localtime(timezone.now()).strftime('%d/%m/%Y à %H:%M')
        canvas.drawString(MARGIN_L, y + 0.45 * cm, f'Exporté le {date_str}')
        centre = "CFLD  ·  Aboisso, Côte d'Ivoire  ·  Document confidentiel"
        cw = canvas.stringWidth(centre, 'Helvetica', 7)
        canvas.drawString((W - cw) / 2, y + 0.45 * cm, centre)
        canvas.setFont('Helvetica-Bold', 7)
        pg = f'Page {doc.page}'
        pw = canvas.stringWidth(pg, 'Helvetica-Bold', 7)
        canvas.drawString(W - MARGIN_R - pw, y + 0.45 * cm, pg)
        canvas.restoreState()

    buffer = io.BytesIO()
    frame = Frame(MARGIN_L, MARGIN_B, CONTENT_W, H - MARGIN_T - MARGIN_B,
                  leftPadding=0, rightPadding=0, topPadding=0, bottomPadding=0)
    template = PageTemplate(id='main', frames=[frame],
                            onPage=lambda c, d: (draw_header(c, d), draw_footer(c, d)))
    doc = BaseDocTemplate(buffer, pagesize=A4, pageTemplates=[template],
                          title=f'Fiche joueur — {joueur.nom_complet}', author='CFLD')

    s_lbl = ParagraphStyle('lbl', fontName='Helvetica', fontSize=7.5, textColor=C_GRIS, leading=10)
    s_val = ParagraphStyle('val', fontName='Helvetica-Bold', fontSize=10, textColor=C_NOIR, leading=13)
    s_sec = ParagraphStyle('sec', fontName='Helvetica-Bold', fontSize=7.5, textColor=C_ROUGE, leading=10)

    def section_title(text):
        p = Paragraph(text.upper(), s_sec)
        t = Table([[p]], colWidths=[CONTENT_W])
        t.setStyle(TableStyle([
            ('LEFTPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 5),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
            ('LINEBEFORE', (0, 0), (0, -1), 3, C_ROUGE),
            ('BACKGROUND', (0, 0), (-1, -1), C_GRIS_LIGHT),
        ]))
        return t

    def fields_grid(pairs, bg=None):
        LABEL_W = 3.2 * cm
        VALUE_W = (CONTENT_W / 2) - LABEL_W
        rows = []
        for i in range(0, len(pairs), 2):
            l1, v1 = pairs[i]
            l2, v2 = pairs[i + 1] if i + 1 < len(pairs) else ('', '')
            rows.append([
                Paragraph(l1, s_lbl), Paragraph(str(v1) if v1 else '—', s_val),
                Paragraph(l2, s_lbl), Paragraph(str(v2) if v2 else '—', s_val),
            ])
        t = Table(rows, colWidths=[LABEL_W, VALUE_W, LABEL_W, VALUE_W])
        style = [
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('TOPPADDING', (0, 0), (-1, -1), 5),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
            ('LEFTPADDING', (0, 0), (-1, -1), 4),
            ('LINEBELOW', (0, 0), (-1, -2), 0.3, C_GRIS_LINE),
            ('LINEBEFORE', (2, 0), (2, -1), 0.3, C_GRIS_LINE),
        ]
        if bg:
            style.append(('BACKGROUND', (0, 0), (-1, -1), bg))
        t.setStyle(TableStyle(style))
        return t

    story = [Spacer(1, 0.2 * cm)]

    photo_cell = Spacer(3.2 * cm, 4.2 * cm)
    if joueur.photo:
        photo_path = os.path.join(settings.MEDIA_ROOT, str(joueur.photo))
        if os.path.exists(photo_path):
            try:
                buf_img = io.BytesIO()
                with PILImage.open(photo_path) as img:
                    if img.mode not in ('RGB', 'L'):
                        img = img.convert('RGB')
                    img.thumbnail((300, 400), PILImage.LANCZOS)
                    img.save(buf_img, format='JPEG', quality=75, optimize=True)
                buf_img.seek(0)
                photo_cell = RLImage(buf_img, width=3.0 * cm, height=3.8 * cm)
            except Exception:
                pass

    photo_wrap = Table([[photo_cell]], colWidths=[3.4 * cm])
    photo_wrap.setStyle(TableStyle([
        ('BOX', (0, 0), (-1, -1), 0.75, C_GRIS_LINE),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('BACKGROUND', (0, 0), (-1, -1), C_GRIS_LIGHT),
    ]))

    FIELDS_W = CONTENT_W - 3.6 * cm
    id_lbl_w = 3.2 * cm
    id_val_w = (FIELDS_W - id_lbl_w) / 2

    ddn = joueur.date_naissance.strftime('%d / %m / %Y') if joueur.date_naissance else '—'
    id_data = [
        [Paragraph('Nom', s_lbl), Paragraph(joueur.nom.upper(), s_val),
         Paragraph('Prénom', s_lbl), Paragraph(joueur.prenom, s_val)],
        [Paragraph('Date de naissance', s_lbl), Paragraph(ddn, s_val),
         Paragraph('Âge', s_lbl), Paragraph(f'{joueur.age} ans', s_val)],
        [Paragraph('Sexe', s_lbl), Paragraph(joueur.get_sexe_display(), s_val),
         Paragraph('Catégorie', s_lbl), Paragraph(joueur.categorie, s_val)],
        [Paragraph('Nationalité', s_lbl), Paragraph(joueur.nationalite or '—', s_val),
         Paragraph('Adresse', s_lbl), Paragraph(joueur.adresse or '—', s_val)],
    ]
    id_table = Table(id_data, colWidths=[id_lbl_w, id_val_w, id_lbl_w, id_val_w])
    id_table.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('LEFTPADDING', (0, 0), (-1, -1), 5),
        ('LINEBELOW', (0, 0), (-1, -2), 0.3, C_GRIS_LINE),
        ('LINEBEFORE', (2, 0), (2, -1), 0.3, C_GRIS_LINE),
        ('ROWBACKGROUNDS', (0, 0), (-1, -1), [C_BLANC, C_ROUGE_LIGHT]),
    ]))

    story.append(section_title('Identité du joueur'))
    story.append(Spacer(1, 0.15 * cm))
    story.append(Table([[photo_wrap, id_table]],
                       colWidths=[3.6 * cm, FIELDS_W],
                       style=[('VALIGN', (0, 0), (-1, -1), 'TOP'),
                              ('LEFTPADDING', (1, 0), (1, 0), 8)]))

    story.append(Spacer(1, 0.35 * cm))
    story.append(section_title('Profil sportif'))
    story.append(Spacer(1, 0.15 * cm))
    story.append(fields_grid([
        ('Poste', joueur.get_poste_display()),
        ('N° de maillot', str(joueur.numero)),
        ('Pied fort', joueur.pied_fort or '—'),
        ('Ancien club', joueur.ancien_club or 'Aucun'),
    ]))

    story.append(Spacer(1, 0.35 * cm))
    story.append(section_title('Statistiques'))
    story.append(Spacer(1, 0.15 * cm))
    story.append(fields_grid([
        ('Matchs joués', str(joueur.matchs_joues)),
        ('Buts', str(joueur.buts)),
        ('N° de licence', joueur.numero_licence or '—'),
        ('Date inscription',
         joueur.date_inscription.strftime('%d/%m/%Y') if joueur.date_inscription else '—'),
    ]))

    if joueur.telephone_whatsapp or joueur.email:
        story.append(Spacer(1, 0.35 * cm))
        story.append(section_title('Coordonnées'))
        story.append(Spacer(1, 0.15 * cm))
        story.append(fields_grid([
            ('WhatsApp', joueur.telephone_whatsapp or '—'),
            ('Email', joueur.email or '—'),
        ]))

    if joueur.nom_parent or joueur.telephone_parent:
        story.append(Spacer(1, 0.35 * cm))
        story.append(section_title('Parent / Tuteur légal'))
        story.append(Spacer(1, 0.15 * cm))
        story.append(fields_grid([
            ('Nom complet', joueur.nom_parent or '—'),
            ('Téléphone', joueur.telephone_parent or '—'),
            ('Email parent', joueur.email_parent or '—'),
            ('', ''),
        ], bg=C_ROUGE_LIGHT))

    extras = []
    if joueur.info_scolaire:
        extras += [('École / Établissement', joueur.info_scolaire)]
    if joueur.contact_urgence:
        extras += [("Contact d'urgence", joueur.contact_urgence)]
    if extras:
        if len(extras) % 2 != 0:
            extras.append(('', ''))
        story.append(Spacer(1, 0.35 * cm))
        story.append(section_title('Informations complémentaires'))
        story.append(Spacer(1, 0.15 * cm))
        story.append(fields_grid(extras))

    doc.build(story)
    pdf_bytes = buffer.getvalue()
    buffer.close()

    fname = f'fiche_{joueur.prenom.lower()}_{joueur.nom.lower()}.pdf'
    response = HttpResponse(pdf_bytes, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{fname}"'
    return response


@staff_required
def admin_joueurs_pdf(request):
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import (
        HRFlowable, Image as RLImage, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle,
    )

    C_ROUGE = colors.HexColor('#C8161D')
    C_ROUGE_DARK = colors.HexColor('#8A0D12')
    C_NOIR = colors.HexColor('#0B0B0B')
    C_GRIS = colors.HexColor('#6B7280')
    C_GRIS_LIGHT = colors.HexColor('#F5F5F5')
    C_GRIS_LINE = colors.HexColor('#E5E7EB')
    C_BLANC = colors.white

    poste = request.GET.get('poste', 'all')
    q = request.GET.get('q', '')
    qs = Joueur.objects.all().order_by('poste', 'numero')
    if q:
        qs = qs.filter(nom__icontains=q) | qs.filter(prenom__icontains=q)
    if poste != 'all':
        qs = qs.filter(poste=poste)
    joueurs = list(qs)

    buffer = io.BytesIO()
    MARGIN = 1.5 * cm
    doc = SimpleDocTemplate(
        buffer, pagesize=landscape(A4),
        leftMargin=MARGIN, rightMargin=MARGIN,
        topMargin=MARGIN, bottomMargin=MARGIN,
        title='Effectif CFLD',
    )

    s_title = ParagraphStyle('t', fontName='Helvetica-Bold', fontSize=18, textColor=C_ROUGE, leading=22)
    s_sub = ParagraphStyle('s', fontName='Helvetica', fontSize=9, textColor=C_GRIS, leading=12)
    s_cell = ParagraphStyle('c', fontName='Helvetica', fontSize=9, textColor=C_NOIR, leading=11)
    s_head = ParagraphStyle('h', fontName='Helvetica-Bold', fontSize=8, textColor=C_BLANC,
                            leading=10, alignment=TA_CENTER)

    story = []
    logo_path = os.path.join(settings.BASE_DIR, 'assets', 'images', 'club', 'logo.png')
    if os.path.exists(logo_path):
        logo = RLImage(logo_path, width=1.4 * cm, height=1.4 * cm)
        hdr_table = Table(
            [[logo, Paragraph('EFFECTIF — CFLD', s_title)]],
            colWidths=[1.6 * cm, None],
        )
        hdr_table.setStyle(TableStyle([
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('LEFTPADDING', (0, 0), (-1, -1), 0),
        ]))
        story.append(hdr_table)
    else:
        story.append(Paragraph('EFFECTIF — CFLD', s_title))

    date_str = timezone.localtime(timezone.now()).strftime('%d/%m/%Y à %H:%M')
    story.append(Paragraph(
        f"Centre de Formation Lancine Diomandé  ·  Exporté le {date_str}  ·  {len(joueurs)} joueur(s)",
        s_sub,
    ))
    story.append(Spacer(1, 0.3 * cm))
    story.append(HRFlowable(width='100%', thickness=1.5, color=C_ROUGE, spaceAfter=6))

    POSTE_LABELS = {'GK': 'Gardien', 'DEF': 'Défenseur', 'MIL': 'Milieu', 'ATT': 'Attaquant'}
    COL_W = [1.0 * cm, 1.5 * cm, 6.5 * cm, 3.0 * cm, 4.5 * cm, 2.2 * cm, 2.2 * cm, 2.5 * cm]
    headers = ['#', 'N°', 'Joueur', 'Poste', 'Nationalité', 'Matchs', 'Buts', 'Statut']
    rows = [[Paragraph(h, s_head) for h in headers]]

    for idx, j in enumerate(joueurs, start=1):
        rows.append([
            Paragraph(str(idx), ParagraphStyle('n', fontName='Helvetica', fontSize=8,
                                               textColor=C_GRIS, alignment=TA_CENTER)),
            Paragraph(str(j.numero) if j.numero else '—',
                      ParagraphStyle('n2', fontName='Helvetica-Bold', fontSize=9,
                                     textColor=C_ROUGE, alignment=TA_CENTER)),
            Paragraph(f'{j.prenom} {j.nom}', s_cell),
            Paragraph(POSTE_LABELS.get(j.poste, j.poste), s_cell),
            Paragraph(j.nationalite or '—', s_cell),
            Paragraph(str(j.matchs_joues), ParagraphStyle('stat', fontName='Helvetica-Bold',
                                                          fontSize=9, alignment=TA_CENTER)),
            Paragraph(str(j.buts), ParagraphStyle('stat2', fontName='Helvetica-Bold',
                                                   fontSize=9, alignment=TA_CENTER)),
            Paragraph('Actif' if j.actif else 'Inactif',
                      ParagraphStyle('st', fontName='Helvetica-Bold', fontSize=8,
                                     textColor=C_ROUGE_DARK if j.actif else C_GRIS,
                                     alignment=TA_CENTER)),
        ])

    table = Table(rows, colWidths=COL_W, repeatRows=1)
    row_bgs = [('BACKGROUND', (0, i), (-1, i), C_BLANC if i % 2 == 1 else C_GRIS_LIGHT)
               for i in range(1, len(rows))]
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), C_ROUGE_DARK),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('LEFTPADDING', (0, 0), (-1, -1), 5),
        ('RIGHTPADDING', (0, 0), (-1, -1), 5),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LINEBELOW', (0, 0), (-1, -1), 0.25, C_GRIS_LINE),
        ('BOX', (0, 0), (-1, -1), 0.5, C_GRIS_LINE),
        *row_bgs,
    ]))
    story.append(table)
    story.append(Spacer(1, 0.4 * cm))
    story.append(HRFlowable(width='100%', thickness=0.5, color=C_GRIS_LINE))
    story.append(Paragraph(
        "CFLD  ·  Aboisso, Côte d'Ivoire  ·  Document confidentiel",
        ParagraphStyle('ft', fontName='Helvetica', fontSize=7, textColor=C_GRIS,
                       alignment=TA_CENTER, leading=10, spaceBefore=4),
    ))

    doc.build(story)
    pdf_bytes = buffer.getvalue()
    buffer.close()

    fname = f'effectif_cfld_{timezone.now().strftime("%Y%m%d")}.pdf'
    response = HttpResponse(pdf_bytes, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{fname}"'
    return response
